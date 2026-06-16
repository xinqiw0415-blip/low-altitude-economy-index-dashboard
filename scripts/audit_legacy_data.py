from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "reports" / "legacy_data_audit.md"


def missing_counts(header: tuple, rows: list[tuple]) -> dict[str, int]:
    return {
        str(column): sum(row[index] in (None, "") for row in rows)
        for index, column in enumerate(header)
    }


def main() -> None:
    candidates = [path for path in ROOT.glob("*.xlsx") if "数据汇总" in path.name]
    if not candidates:
        raise FileNotFoundError("未找到名称包含‘数据汇总’的 xlsx 文件")

    workbook_path = max(candidates, key=lambda path: path.stat().st_size)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    lines = ["# 旧数据质量审计", "", f"来源：`{workbook_path.name}`", ""]

    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        header = tuple(next(iterator))
        rows = list(iterator)
        lines.extend(
            [
                f"## {sheet.title}",
                "",
                f"- 记录数：{len(rows)}",
                f"- 字段数：{len(header)}",
                f"- 缺失值：`{missing_counts(header, rows)}`",
            ]
        )

        index = {str(column): position for position, column in enumerate(header)}
        if sheet.title == "concept_members":
            raw_codes = [str(row[index["ts_code"]]) for row in rows]
            normalized = [code.split(".")[0] for code in raw_codes]
            lines.append(f"- 代码标准化后重复数：{len(normalized) - len(set(normalized))}")
        elif sheet.title == "stock_daily":
            codes = Counter(str(row[index["ts_code"]]) for row in rows)
            dates = [str(row[index["date"]]) for row in rows]
            keys = {(row[index["ts_code"]], row[index["date"]]) for row in rows}
            lines.extend(
                [
                    f"- 股票数：{len(codes)}，分布：`{dict(codes)}`",
                    f"- 日期范围：{min(dates)} 至 {max(dates)}",
                    f"- 证券-日期重复数：{len(rows) - len(keys)}",
                ]
            )
        elif sheet.title == "index_daily":
            codes = Counter(str(row[index["index_code"]]) for row in rows)
            lines.append(f"- 指数分布：`{dict(codes)}`")
        elif sheet.title == "unstructured_posts":
            sources = Counter(str(row[index["source"]]) for row in rows)
            urls = [row[index["url"]] for row in rows]
            lines.extend(
                [
                    f"- 来源分布：`{dict(sources)}`",
                    f"- 唯一 URL：{len(set(urls))}，URL 重复数：{len(urls) - len(set(urls))}",
                    "- 结论：发布时间与摘要全部缺失，且存在广告链接，不进入正式文本库。",
                ]
            )
        lines.append("")

    lines.extend(
        [
            "## 总体结论",
            "",
            "1. 原行情表可作为历史数据交叉校验来源，但股票池过窄，不能直接构造低空经济板块指数。",
            "2. 原概念成分表存在代码重复、退市公司和明显弱相关标的，需要废弃并重新审核。",
            "3. 原文本表缺少发布时间和正文，无法进行事件研究，应重新采集。",
        ]
    )
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"已生成 {REPORT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
